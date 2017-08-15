# coding=utf-8
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


def list2xls(objects, attributes, path, append=False):
    """
    把objects的所有attributes指定的属性保存在path指定的xlsx中
    :param objects: list 对象列表
    :param attributes: list[str] 属性名称列表
    :param path: str 保存路径
    :param append: bool 是否追加 
    """
    if append:
        wb = load_workbook(path)
        ws = wb.active
        if ws.title == 'sheet1' and objects:
            ws.title = objects[0].__class__.__name__
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = objects[0].__class__.__name__ if objects else 'sheet1'
        ws.append(attributes)
    for o in objects:
        ws.append([getattr(o, a, '') for a in attributes])
    if objects:
        tab = Table(displayName=ws.title, ref="%c%d:%c%d" % (ord('A'), 1, ord('@') + ws.max_column, ws.max_row))
        style = TableStyleInfo(name='TableStyleMedium9', showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
    wb.save(path)
